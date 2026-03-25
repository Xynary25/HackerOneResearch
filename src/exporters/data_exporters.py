import json
import csv
import logging
from pathlib import Path
from typing import List, Dict, Any, Optional
from datetime import datetime

# ✅ Проверка наличия openpyxl для Excel
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo

    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    logging.warning("⚠ openpyxl не установлен — Excel экспорт недоступен")
    logging.warning("💡 Установите: pip install openpyxl")

logger = logging.getLogger(__name__)


class JSONExporter:
    """
    Экспорт данных в JSON формат

    ✅ Преимущества:
    - Сохраняет типы данных
    - Поддерживает вложенные структуры
    - Читаемый формат

    ✅ Используется для:
    - Дальнейшей обработки
    - Резервных копий
    - Интеграций
    """

    def __init__(self, output_dir: Path):
        self.output_dir = output_dir
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def export(self, data: List[Dict], filename: str) -> Optional[str]:
        """
        Экспорт данных в JSON файл

        Args:
            data: Список словарей для экспорта
            filename: Имя файла (без расширения)

        Returns:
            Путь к файлу или None если ошибка
        """
        if not data:
            logger.warning(f"⚠ Нет данных для экспорта JSON: {filename}")
            return None

        try:
            filepath = self.output_dir / f"{filename}.json"

            with open(filepath, 'w', encoding='utf-8') as f:
                json.dump(
                    data,
                    f,
                    indent=2,
                    ensure_ascii=False,
                    default=str
                )

            logger.info(f"📁 Экспорт JSON: {filepath} ({len(data)} записей)")
            return str(filepath)

        except Exception as e:
            logger.error(f"❌ Ошибка экспорта JSON {filename}: {e}")
            return None


class CSVExporter:
    """
    Экспорт данных в CSV формат

    ✅ Преимущества:
    - Открывается в Excel
    - Легковесный формат
    - Поддерживается везде

    ✅ Используется для:
    - Быстрого просмотра
    - Импорта в другие системы
    - Отчётности
    """

    def __init__(self, output_dir: Path):
        self.output_dir = output_dir
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def export(self, data: List[Dict], filename: str) -> Optional[str]:
        """
        Экспорт данных в CSV файл

        Args:
            data: Список словарей для экспорта
            filename: Имя файла (без расширения)

        Returns:
            Путь к файлу или None если ошибка
        """
        if not data:
            logger.warning(f"⚠ Нет данных для экспорта CSV: {filename}")
            return None

        try:
            filepath = self.output_dir / f"{filename}.csv"

            # ✅ Получаем все ключи из всех записей
            fieldnames = list(data[0].keys())

            with open(filepath, 'w', newline='', encoding='utf-8-sig') as f:
                writer = csv.DictWriter(
                    f,
                    fieldnames=fieldnames,
                    quoting=csv.QUOTE_MINIMAL
                )
                writer.writeheader()
                writer.writerows(data)

            logger.info(f"📁 Экспорт CSV: {filepath} ({len(data)} записей)")
            return str(filepath)

        except Exception as e:
            logger.error(f"❌ Ошибка экспорта CSV {filename}: {e}")
            return None


class ExcelExporter:
    """
    Экспорт данных в Excel формат (XLSX)

    ✅ Преимущества:
    - Форматирование ячеек
    - Авто-ширина колонок
    - Таблицы с фильтрами
    - Визуальная привлекательность

    ✅ Используется для:
    - Презентаций
    - Отчётности для руководства
    - Совместной работы

    ⚠️ Требует: pip install openpyxl
    """

    def __init__(self, output_dir: Path):
        self.output_dir = output_dir
        self.output_dir.mkdir(parents=True, exist_ok=True)

        # ✅ Стили для форматирования
        self.header_font = Font(bold=True, color="FFFFFF", size=12)
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.header_alignment = Alignment(horizontal="center", vertical="center")

        self.cell_font = Font(size=11)
        self.cell_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    def export(self, data: List[Dict], filename: str) -> Optional[str]:
        """
        Экспорт данных в Excel файл с форматированием

        Args:
            data: Список словарей для экспорта
            filename: Имя файла (без расширения)

        Returns:
            Путь к файлу или None если ошибка
        """
        if not EXCEL_AVAILABLE:
            logger.error("❌ Excel экспорт недоступен — установите openpyxl")
            logger.error("💡 Команда: pip install openpyxl")
            return None

        if not data:
            logger.warning(f"⚠ Нет данных для экспорта Excel: {filename}")
            return None

        try:
            filepath = self.output_dir / f"{filename}.xlsx"

            # ✅ Создаём workbook и worksheet
            wb = Workbook()
            ws = wb.active
            ws.title = filename[:31]  # Excel ограничивает имя 31 символом

            # ✅ Получаем заголовки
            headers = list(data[0].keys())

            # ✅ Записываем заголовки
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.alignment = self.header_alignment
                cell.border = self.border

            # ✅ Записываем данные
            for row_num, row_data in enumerate(data, 2):
                for col_num, header in enumerate(headers, 1):
                    value = row_data.get(header, "")

                    # ✅ Конвертируем сложные типы
                    if isinstance(value, (list, dict)):
                        value = str(value)
                    elif value is None:
                        value = ""

                    cell = ws.cell(row=row_num, column=col_num, value=value)
                    cell.font = self.cell_font
                    cell.alignment = self.cell_alignment
                    cell.border = self.border

                    # ✅ Авто-высота строки для длинного текста
                    if isinstance(value, str) and len(value) > 50:
                        ws.row_dimensions[row_num].height = 40

            # ✅ Авто-ширина колонок
            for col_num, header in enumerate(headers, 1):
                max_length = len(str(header))
                for row in data:
                    cell_value = row.get(header, "")
                    if cell_value:
                        max_length = max(max_length, len(str(cell_value)))

                # ✅ Ограничиваем ширину (макс 50 символов)
                adjusted_width = min(max_length + 2, 50)
                column_letter = get_column_letter(col_num)
                ws.column_dimensions[column_letter].width = adjusted_width

            # ✅ Добавляем таблицу с фильтрами
            table_ref = f"A1:{get_column_letter(len(headers))}{len(data) + 1}"
            table = Table(displayName=f"Table_{filename[:20]}", ref=table_ref)
            style = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False
            )
            table.tableStyleInfo = style
            ws.add_table(table)

            # ✅ Замораживаем первую строку
            ws.freeze_panes = "A2"

            # ✅ Сохраняем файл
            wb.save(filepath)
            wb.close()

            logger.info(f"📁 Экспорт Excel: {filepath} ({len(data)} записей)")
            return str(filepath)

        except Exception as e:
            logger.error(f"❌ Ошибка экспорта Excel {filename}: {e}")
            import traceback
            logger.error(traceback.format_exc())
            return None


class ExportManager:
    """
    Менеджер экспорта — координирует экспорт во все форматы

    ✅ Преимущества:
    - Единый интерфейс для всех форматов
    - Автоматическое создание директорий
    - Удаление дубликатов
    - Логирование всех операций

    ✅ Используется в:
    - main.py
    - gui.py
    """

    def __init__(self, output_dir: Path, export_excel: bool = False):
        """
        Инициализация менеджера экспорта

        Args:
            output_dir: Директория для сохранения файлов
            export_excel: Включить экспорт в Excel
        """
        self.output_dir = output_dir
        self.output_dir.mkdir(parents=True, exist_ok=True)

        self.json_exporter = JSONExporter(output_dir)
        self.csv_exporter = CSVExporter(output_dir)
        self.excel_exporter = ExcelExporter(output_dir) if export_excel else None
        self.exported_files = []

        logger.info(f"📂 ExportManager инициализирован: {output_dir}")
        logger.info(f"📊 Excel экспорт: {'включён' if export_excel else 'выключен'}")

    def _remove_duplicates(self, data: List[Dict], key: str) -> List[Dict]:
        """
        Удаление дубликатов по ключу

        Args:
            data: Список словарей
            key: Ключ для проверки уникальности

        Returns:
            Список без дубликатов
        """
        seen = set()
        unique_data = []

        for item in data:
            identifier = item.get(key)
            if identifier and identifier not in seen:
                seen.add(identifier)
                unique_data.append(item)

        removed_count = len(data) - len(unique_data)
        if removed_count > 0:
            logger.info(f"🗑 Удалено {removed_count} дубликатов по ключу '{key}'")

        return unique_data

    def export_all(
            self,
            hackers: List[Dict],
            analyses: List[Dict],
            reports: List[Dict],
            formats: List[str] = None
    ) -> List[str]:
        """
        Экспорт всех данных в указанные форматы

        Args:
            hackers: Данные хакеров
            analyses: Данные анализа
            reports: Данные отчётов
            formats: Список форматов ['json', 'csv', 'excel']

        Returns:
            Список путей к экспортированным файлам
        """
        formats = formats or ["json", "csv"]
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

        logger.info(f"📤 Начало экспорта: форматы={formats}")

        # ✅ Удаляем дубликаты
        hackers = self._remove_duplicates(hackers, 'username')
        analyses = self._remove_duplicates(analyses, 'username')
        reports = self._remove_duplicates(reports, 'report_id')

        # ✅ Экспорт в JSON
        if "json" in formats:
            if hackers:
                result = self.json_exporter.export(hackers, f"hackers_{timestamp}")
                if result:
                    self.exported_files.append(result)

            if analyses:
                result = self.json_exporter.export(analyses, f"analyses_{timestamp}")
                if result:
                    self.exported_files.append(result)

            if reports:
                result = self.json_exporter.export(reports, f"reports_{timestamp}")
                if result:
                    self.exported_files.append(result)

        # ✅ Экспорт в CSV
        if "csv" in formats:
            if hackers:
                result = self.csv_exporter.export(hackers, f"hackers_{timestamp}")
                if result:
                    self.exported_files.append(result)

            if analyses:
                result = self.csv_exporter.export(analyses, f"analyses_{timestamp}")
                if result:
                    self.exported_files.append(result)

            if reports:
                result = self.csv_exporter.export(reports, f"reports_{timestamp}")
                if result:
                    self.exported_files.append(result)

        # ✅ Экспорт в Excel
        if "excel" in formats and self.excel_exporter:
            if hackers:
                result = self.excel_exporter.export(hackers, f"hackers_{timestamp}")
                if result:
                    self.exported_files.append(result)

            if analyses:
                result = self.excel_exporter.export(analyses, f"analyses_{timestamp}")
                if result:
                    self.exported_files.append(result)

        logger.info(f"✅ Экспорт завершён: {len(self.exported_files)} файлов")

        return self.exported_files

    def get_exported_files(self) -> List[str]:
        """Получить список всех экспортированных файлов"""
        return self.exported_files.copy()

    def clear(self):
        """Очистить список экспортированных файлов"""
        self.exported_files = []
        logger.info("🗑 Список экспортированных файлов очищен")